import os
import zipfile
from os.path import basename
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook

path_to_the_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), '../file')
path_to_the_created_archive = os.path.join(os.path.dirname(os.path.abspath(__file__)), '../resources')
path_zip = os.path.join(path_to_the_created_archive, "created_zip_for_test.zip")
list_files = os.listdir(path_to_the_files)


def test_create_arhive(cleaning_the_directory_before_the_test):
    with zipfile.ZipFile(path_zip, mode='w', compression=zipfile.ZIP_STORED) as arch:
        for f in list_files:
            add_file = os.path.join(path_to_the_files, f)
            arch.write(add_file, basename(add_file))
    file = os.listdir(path_to_the_created_archive)
    assert len(file) == 1, f"Неверное количество созданных архивов {len(file)} не ровно {1}"


def test_for_pdf_format():
    with zipfile.ZipFile(path_zip) as arhv:
        pdf_file = arhv.extract("pdf_file.pdf")
        read = PdfReader(pdf_file)
        page = read.pages[0]
        text = page.extract_text()
    assert "Это документ в формате pdf" in text
    arhv.close()




def test_for_csv_format():
    with zipfile.ZipFile(path_zip) as arhv:
        csv_ = arhv.extract("csv_file.csv")
        with open(csv_) as csvfile:
            csvfile = csv.reader(csvfile)
            list_csv = []
            for r in csvfile:
                text = "".join(r).replace(";", " ")
                list_csv.append(text)
            assert list_csv[0] == "0 вывод текста тестового ", f"В файле отсутствует информация"
        os.remove(csv_)
        arhv.close()


def test_for_xlsx_format():
    with zipfile.ZipFile(path_zip) as arhv:
        xlsx_ = arhv.extract("xlsx_file.xlsx", path=path_to_the_created_archive)
        with open(xlsx_):
            xlsxfile = load_workbook(xlsx_)
            sheet = xlsxfile.active
            assert sheet.cell(row=4, column=2).value == "В формате xlsx", f"Фактический результат {sheet.cell(row=4, column=2).value}"
        os.remove(xlsx_)
        arhv.close()



